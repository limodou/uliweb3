from ...utils._compat import iteritems, text_type, PY2, u, callable

def get_model_tables(tables, appname):
    t = []
    for tablename, m in iteritems(tables):
        if hasattr(m, '__appname__') and m.__appname__ == appname:
            t.append(tablename)
    return t

def safe_str(s, encoding='utf-8'):
    if PY2 and isinstance(s, six.text_type):
        return s.encode(encoding)
    else:
        return u(s, encoding)


def generate_excel(tables, apps, filename, **kwargs):
    from uliweb import orm
    from os.path import dirname, join    
    from uliweb.orm import ReferenceProperty
    from sqlalchemy.schema import CreateIndex
    
    menus = []
    table_objs = {}
    for app in apps:
        section = {
            'name': app,
            'items': []
        }

        t = get_model_tables(tables, app)
        if not t: continue
        for tablename in t:       
            item = {
                'app_name': app.replace('.', '_'),
                'name': tablename,
                'caption': tablename,
            }
            table_objs[tablename] = item
            try:
                M = orm.get_model(tablename)
            except:
                continue     
            
            item['label'] = getattr(M, '__verbose_name__', tablename)
            if tablename != M.tablename:
                item['caption'] += ' - ' + M.tablename
            
            section['items'].append(item)
        menus.append(section)
    
    
    all_tables = []
    for name, t in sorted(iteritems(tables)):
        model = {
            'name': name,
            'fields': [],
            'relations': [],
            'choices': [],
            'indexes': [],
        }
        if hasattr(t, '__appname__'):
            model['appname'] = t.__appname__
        else:
            model['appname'] = None
        
        M = None
        fields = t.c.keys()
        try:
            M = orm.get_model(name)
            fields = [x[0] for x in M._fields_list if x[0] in t.c]
        except:
            pass

        if 'id' in fields:
            fields.remove('id')
            fields.insert(0, 'id')
    
        if getattr(M, '__verbose_name__', None):
            model['label'] = "%s(%s)" % (name, getattr(M, '__verbose_name__', None))
        else:
            model['label'] = name
        if name != getattr(M, 'tablename', name):
            model['label'] += ' - ' + M.tablename
            
        #Add docstring for Model
        if M.__doc__:
            model['desc'] = M.__doc__
        else:
            model['desc'] = ''
        table = table_objs[name]
        table['desc'] = model['desc']
        
        #process indexes
        for x in t.indexes:
            model['indexes'].append(CreateIndex(x))
        
        star_index = 0
        for key in fields:
            tablefield = t.c[key]
            field = {
                'name': tablefield.name,
                'type': tablefield.type,
                'nullable': tablefield.nullable,
                'primary_key': tablefield.primary_key
            }
            field['reftable'] = None
            field['choices_name'] = ''
            field['star'] = False
            field['label'] = tablefield.name
            
            if M:
                ppp = M.properties[tablefield.name]
                if getattr(ppp, 'verbose_name', None):
                    field['label'] = (getattr(ppp, 'verbose_name', None))

                if getattr(ppp, 'choices_name'):
                    field['choices_name'] = getattr(ppp, 'choices_name')
                
                if getattr(ppp, 'choices', None):
                    choices_list = getattr(ppp, 'choices', None)
                    if callable(choices_list) :
                        choices_list = choices_list()
                    if choices_list :
                        star_index = star_index + 1
                        model['choices'].append({
                            'index': star_index,
                            'fieldlabel': field['label'],
                            'fieldname': field['name'],
                            'list': choices_list})
                        field['star'] = star_index
    
                if ppp and ppp.__class__ is ReferenceProperty:
                    field['reftable'] = ppp.reference_class.tablename


            model['fields'].append(field)
        all_tables.append(model)
    database = {}
    database["menus"] = menus
    database["tables"] = all_tables
    return write_excel(filename, database)
    
def write_excel(filename, database):
    import openpyxl

    wb = openpyxl.Workbook(write_only=True)
    sh = wb.create_sheet(title="表索引")
    sh.append(['模块名', '表名', '中文名', '说明'])
    for menu in database['menus']:
        for item in menu['items']:
            sh.append([item['app_name'], item['name'], item['label'], item['desc']])
    app = ''
    for table in sorted(database['tables'], key=lambda x: x['appname']):
        if table['appname'] != app:
            app = table['appname']
            sh = wb.create_sheet(title=app)

        sh.append(['表名', table['name'], '中文名', table['label']])
        sh.append(['中文名', '字段名', '类型', '是否允许为空', '是否主键', '关系', '参数名'])
        for field in table['fields']:
            sh.append([str(field['label']), field['name'], str(field['type']),
                       'Y' if field['nullable'] else 'N',
                       'Y' if field['primary_key'] else '', field['reftable'],
                       field['choices_name']])

        sh.append([''])
        if table['indexes']:
            sh.append(['索引：'])
            for index in table['indexes']:
                sh.append([str(index)])
            sh.append([''])
            sh.append([''])

    wb.save(filename)
